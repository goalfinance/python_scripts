import openpyxl
import os
from goalfinance.utils.calendar import month_to_str
import numpy as np
import re

def load_workbook(file_path):
    if file_path == None:
        return None
    if (os.path.exists(file_path)):
        return openpyxl.load_workbook(file_path)
    else:
        return None

#Check if the sheet for the month exists
def does_monsheet_exist(workbook, mon):
    month_str = month_to_str(mon)
    sheet_names = workbook.sheetnames
    for sheet_name in sheet_names:
        if sheet_name == month_str:
            return True
    
    return False

def get_month_sheet(workbook, mon):
    month_str = month_to_str(mon)
    if does_monsheet_exist(workbook, mon) == True:
        return workbook[month_str]
    else:
        return None

def source_attendance_group_by_member(source_workbook, month):
    month_sheet = get_month_sheet(source_workbook, month)
    if month_sheet == None:
        return None
    source_attendances = dict()
    i = 0
    for row in month_sheet.rows:
        if i == 0:
            i += 1
            continue
        member_name = row[2].value
        members_attendances_records = []
        if member_name not in source_attendances:
            source_attendances[member_name] = []
            members_attendances_records = source_attendances[member_name]
        else:
            members_attendances_records = source_attendances[member_name]
            if members_attendances_records == None:
                members_attendances_records = []
                source_attendances[member_name] = members_attendances_records

        members_attendances_records.append(row)
        
    
    attendances_matrix = dict()

    for member_name in list(source_attendances):
        row_cnt = 0
        column_cnt = month_sheet.max_column
        if source_attendances[member_name] != None:
            row_cnt = len(source_attendances[member_name])
            if row_cnt > 0:
                matrix = np.zeros((row_cnt, 31))
                attendances_matrix[member_name] = matrix
                x = 0
                for row in source_attendances[member_name]:
                    y = 0
                    for column_number in range(3, 34):
                        cell = row[column_number]
                        is_merged_cell = type(cell) is openpyxl.cell.cell.MergedCell
                        if is_merged_cell == False and cell.value != None and re.search("^absen[a-z]*$", cell.value, re.IGNORECASE) != None:
                            matrix[x, y] = 100
                        elif is_merged_cell == False and cell.value != None and re.search("^absen[a-z]*[ ]* half day", cell.value, re.IGNORECASE) != None:
                            matrix[x, y] = 50
                        elif is_merged_cell or cell.value != None:
                            matrix[x, y] = 1
                       
                        y += 1

                    x += 1
    
    return attendances_matrix


           

