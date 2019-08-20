import openpyxl
import os
from goalfinance.utils.calendar import (month_to_str, get_calendar_of_month)
import numpy as np
import re
from calendar import monthrange
from goalfinance.utils.utils import Const


table_title_const = Const()
table_title_const.EXCEL_TABLE_HEADER_MEMBER_NAME = 'Name'
table_title_const.EXCEL_TABLE_HEADER_ATTENDANCE_DAYS = 'Attentance days of month'

def load_workbook(file_path):
    if file_path == None:
        return None
    if (os.path.exists(file_path)):
        return openpyxl.load_workbook(file_path)
    else:
        return None

#Check if the sheet for the month exists
def does_monsheet_exist(workbook, mon):
    month_str = month_to_str(mon)
    sheet_names = workbook.sheetnames
    for sheet_name in sheet_names:
        if sheet_name == month_str:
            return True
    
    return False

def get_month_sheet(workbook, mon):
    month_str = month_to_str(mon)
    if does_monsheet_exist(workbook, mon) == True:
        return workbook[month_str]
    else:
        return None

def get_attendances_matrix(source_workbook, year, month):
    month_sheet = get_month_sheet(source_workbook, month)
    if month_sheet == None:
        return None
    source_attendances = dict()
    i = 0
    for row in month_sheet.rows:
        if i == 0:
            i += 1
            continue
        member_id = row[2].value
        if member_id == None:
            continue
        members_attendances_records = []
        if member_id not in source_attendances:
            source_attendances[member_id] = []
            members_attendances_records = source_attendances[member_id]
        else:
            members_attendances_records = source_attendances[member_id]
            if members_attendances_records == None:
                members_attendances_records = []
                source_attendances[member_id] = members_attendances_records

        members_attendances_records.append(row)
        
    
    attendances_matrix = dict()
    members_full_name = dict()
    days_of_month = monthrange(year, month)[1]
    """
    Group the attendances information by member
    """
    for member_id in list(source_attendances):
        row_cnt = 0       
        if source_attendances[member_id] != None:
            row_cnt = len(source_attendances[member_id])
            if row_cnt > 0:
                matrix = np.zeros((row_cnt, days_of_month))
                attendances_matrix[member_id] = matrix
                x = 0
                for row in source_attendances[member_id]:
                    y = 0
                    if member_id not in members_full_name and row[1].value != None:
                        members_full_name[member_id] = row[1].value
                    for column_number in range(3, days_of_month + 3):
                        cell = row[column_number]
                        is_merged_cell = type(cell) is openpyxl.cell.cell.MergedCell
                        if is_merged_cell == False and cell.value != None and re.search("^absen[a-z]*$", cell.value, re.IGNORECASE) != None:
                            matrix[x, y] = 100
                        elif is_merged_cell == False and cell.value != None and re.search("^absen[a-z]*[ ]*half day", cell.value, re.IGNORECASE) != None:
                            matrix[x, y] = 50
                        elif is_merged_cell or cell.value != None:
                            matrix[x, y] = 1
                       
                        y += 1

                    x += 1
    """
    Combine the attendances information of each member into one record.
    """
    attendances_matrix_compressed = dict()
    for member_id in list(attendances_matrix):
        attendances_matrix_per_member_compressed = np.zeros(days_of_month)
        attendances_matrix_compressed[member_id] = attendances_matrix_per_member_compressed
        attendances_matrix_per_member = attendances_matrix[member_id]
        
        attendances_matrix_per_member_transposed = attendances_matrix_per_member.T
        for i in range(0, days_of_month):
            attendances_matrix_per_member_compressed[i] = np.max(attendances_matrix_per_member_transposed[i])
            
    return attendances_matrix_compressed, members_full_name

def create_table_header(worksheet, year, month):
    calendar_of_month = get_calendar_of_month(year, month)
    headers = [table_title_const.EXCEL_TABLE_HEADER_MEMBER_NAME, table_title_const.EXCEL_TABLE_HEADER_ATTENDANCE_DAYS]
    for calendar_info in calendar_of_month:
        headers.append(calendar_info[1])
    
    worksheet.append(headers)

def fill_attendances_matrix_to_target_worksheet(worksheet, attendances_matrix, members_full_name, year, month):
    calendar_of_month = get_calendar_of_month(year, month)
    for member_id in list(attendances_matrix):
        member_full_name = members_full_name[member_id]
        
        attendance_matrix = attendances_matrix[member_id]
        if attendance_matrix != None and np.size(attendance_matrix) == len(calendar_of_month):
            if member_full_name == None:
                member_full_name = member_id
            attendance_record = [member_full_name, 0]

            for calendar_info in calendar_of_month:
                attendance_value = attendance_matrix[calendar_info[1] - 1]
                if attendance_value == 100:
                    attendance_record.append(0)
                elif attendance_value == 50:
                    attendance_record.append(0.5)
                else:
                    attendance_record.append(1)

            worksheet.append(attendance_record)

            #put the formula for calculating the attendance of the month to the cell
            first_day_cell_coordinate = worksheet.cell(row=worksheet._current_row, column=3).coordinate
            last_day_cel_coordinate = worksheet.cell(row=worksheet._current_row, column=worksheet.max_column).coordinate
            attendance_of_month_cell = worksheet.cell(row=worksheet._current_row, column=2)
            formula_for_attendance = "=SUM(" + first_day_cell_coordinate + ":" + last_day_cel_coordinate + ")"
            attendance_of_month_cell.value = formula_for_attendance


