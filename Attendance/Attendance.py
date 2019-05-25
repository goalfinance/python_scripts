import sys, getopt
import time
import os
import openpyxl
from enum import Enum
import re
import calendar

class Command(Enum):
    addMember = 1
    initial = 2
    recordLeaving = 3

class _const:
    class ConstError(TypeError):pass
    def __setattr__(self, name, value):
        if name in self.__dict__:
            raise self.ConstError("Can't rebind const (%s)" %name)
        self.__dict__[name] = value
class MissingParamsError(ValueError):
    pass

class IllegalParamsError(ValueError):
    pass

class MemberNotFoundError(ValueError):
    pass

class IncorrectOperationError(ValueError):
    pass

class ExcelFileOperationError(ValueError):
    pass

const = _const()
const.APP_PARAMS_MEMBER_NAME = 'member_name'
const.APP_PARAMS_LEAVING_DATE = 'leaving_date'
const.APP_PARAMS_FILE_NAME = 'file_name'
const.APP_PARAMS_INITIAL_DATE = 'initial_date'
const.EXCEL_WORKBOOK_MEMBER_SHEET_NAME = 'member list'
const.EXCEL_TABLE_HEADER_MEMBER_NAME = 'Name'
const.EXCEL_TABLE_HEADER_ATTENDANCE_DAYS = 'Attentance days of month'

app_params = dict(member_name="", leaving_date="", file_name="",initial_date="")

def month_to_str(month):
    month_str = str(month)
    if len(month_str) < 2:
        month_str = '0' + month_str 
    return month_str

#Check if the sheet for the month exists
def does_monsheet_exist(workbook, mon):
    month_str = month_to_str(mon)
    sheet_names = workbook.sheetnames
    for sheet_name in sheet_names:
        if sheet_name == month_str:
            return True
    
    return False

def get_month_sheet(workbook, mon):
    month_str = month_to_str(mon)
    if does_monsheet_exist(workbook, mon) == True:
        return workbook[month_str]
    else:
        return None

def get_current_month_num():
    localtime = time.localtime(time.time())
    return localtime.tm_mon

def get_previous_month_num():
    current_month_num = get_current_month_num()
    previous_month_num = current_month_num - 1
    if previous_month_num < 1 :
        previous_month_num = 12
    return previous_month_num

def validate_param_memeber_name(app_params):
    member_name = app_params[const.APP_PARAMS_MEMBER_NAME]
    if member_name == None or member_name == '':
        raise MissingParamsError("The name of member you want to add is missing, please assign it by using '-m' or '--member_name'")
    return member_name

def validate_param_initial_date(app_params):
    initial_date = None
    if app_params[const.APP_PARAMS_INITIAL_DATE] != None and app_params[const.APP_PARAMS_INITIAL_DATE] != "":
        try:
            initial_date = time.strptime(app_params[const.APP_PARAMS_INITIAL_DATE], "%Y-%m")
            return initial_date
        except ValueError as err:
            raise IllegalParamsError("The value of '" + const.APP_PARAMS_INITIAL_DATE + "' is illegal, the correct format is '%Y-%m', for example, '2019-05'", err)
    else:
        raise MissingParamsError("Initial date is missing, please assign it by using '-d' or '--initial_date'")

def validate_param_leaving_date(app_params):
    leaving_date = None
    if app_params[const.APP_PARAMS_LEAVING_DATE] != None and app_params[const.APP_PARAMS_LEAVING_DATE] != "":
        try:
            leaving_date = time.strptime(app_params[const.APP_PARAMS_LEAVING_DATE], "%Y-%m-%d")
            return leaving_date
        except ValueError as err:
            raise IllegalParamsError("The value of '" + const.APP_PARAMS_LEAVING_DATE + "' is illegal, the correct format is '%Y-%m-%d', for example, '2019-05-15'", err)
    else:
        raise MissingParamsError("Leaving date is missing, please assign it by using '-l' or '--leaving_date'")    
            

def perform_add_member(workbook, app_params):
    member_name = validate_param_memeber_name(app_params)
    try:
        member_list_sheet = workbook.get_sheet_by_name(const.EXCEL_WORKBOOK_MEMBER_SHEET_NAME)
    except KeyError:
        member_list_sheet = workbook.create_sheet(const.EXCEL_WORKBOOK_MEMBER_SHEET_NAME, index=0)
    
    new_row = [member_name]
    member_list_sheet.append(new_row)
    workbook.save(app_params[const.APP_PARAMS_FILE_NAME])

def does_member_exist(workbook, member_name):
    try:
        member_list_sheet = workbook[const.EXCEL_WORKBOOK_MEMBER_SHEET_NAME]
        for row in member_list_sheet.rows:
            if re.search(member_name, row[0].value, re.IGNORECASE) != None:
                return True
        return True
    except KeyError:
        return False

"""
Return a list of a tuple, the structure of the tuple is (weekday, day, isHoliday)
"""
def get_calendar_of_month(year, month):
    calendar_of_month = []
    days_of_month = calendar.monthrange(year, month)[1]
    for i in range(1, days_of_month + 1):
        calendar_info = ()
        weekday = calendar.weekday(year, month, i)
        isHoliday = False
        if weekday in [5, 6]:
            isHoliday = True
        calendar_info = weekday, i, isHoliday
        calendar_of_month.append(calendar_info)
    return calendar_of_month

def create_table_header(worksheet, year, month):
    calendar_of_month = get_calendar_of_month(year, month)
    headers = [const.EXCEL_TABLE_HEADER_MEMBER_NAME, const.EXCEL_TABLE_HEADER_ATTENDANCE_DAYS]
    for calendar_info in calendar_of_month:
        headers.append(calendar_info[1])
    
    worksheet.append(headers)

def insert_new_attendance_info(worksheet, member_name, year, month):
    calendar_of_month = get_calendar_of_month(year, month)
    attendance_record = [member_name, 0]
    for calendar_info in calendar_of_month:
        #if it's holiday
        if calendar_info[2] == True:
            attendance_record.append(0)
        else:
            attendance_record.append(1)
    worksheet.append(attendance_record)

    #put the formula for calculating the attendance of the month to the cell
    first_day_cell_coordinate = worksheet.cell(row=worksheet._current_row, column=3).coordinate
    last_day_cel_coordinate = worksheet.cell(row=worksheet._current_row, column=worksheet.max_column).coordinate
    attendance_of_month_cell = worksheet.cell(row=worksheet._current_row, column=2)
    formula_for_attendance = "=SUM(" + first_day_cell_coordinate + ":" + last_day_cel_coordinate + ")"
    attendance_of_month_cell.value = formula_for_attendance
    
def perform_initial_attendance(workbook, app_params):
    member_name = validate_param_memeber_name(app_params)
    initial_date = validate_param_initial_date(app_params)
    year = initial_date.tm_year
    month = initial_date.tm_mon

    month_sheet = get_month_sheet(workbook, month)
    if month_sheet == None:
        month_sheet = workbook.create_sheet(month_to_str(month)) 
    
    if month_sheet.max_row <= 1:
        create_table_header(month_sheet, year, month)

    if does_member_exist(workbook, member_name) == False:
        raise MemberNotFoundError("The member whose name is '" + member_name + "' doesn't exist in the member list, you can add member by option '-a'")
    
    for row in month_sheet.rows:
        if re.search(member_name, row[0].value, re.IGNORECASE) != None:
            return

    insert_new_attendance_info(month_sheet, member_name, year, month)

    workbook.save(app_params[const.APP_PARAMS_FILE_NAME])

def perform_record_leaving(workbook, app_params):
    member_name = validate_param_memeber_name(app_params)
    leaving_date = validate_param_leaving_date(app_params)
    month = leaving_date.tm_mon
    month_day = leaving_date.tm_mday
    mday_column_index = month_day + 1

    month_sheet = get_month_sheet(workbook, month)
    if month_sheet == None or month_sheet.max_row <= 1:
        raise IncorrectOperationError("Incorrect operation, you need do the initial operation first.")

    if does_member_exist(workbook, member_name) == False:
        raise MemberNotFoundError("The member whose name is '" + member_name + "' doesn't exist in the member list, you can add member by option '-a'")
    
    recording_flag = False
    for row in month_sheet.rows:
        if re.search(member_name, row[0].value, re.IGNORECASE) != None:
            row[mday_column_index].value = 0
            recording_flag = True
    if recording_flag == False:
        raise IncorrectOperationError("Incorrect operation, did not find the member[" + member_name + "]'s attendance, need to initial it first.")
    else:
        workbook.save(app_params[const.APP_PARAMS_FILE_NAME])

def main():
    try:
        opts, args = getopt.getopt(sys.argv[1:], "airm:l:f:d:", ["member_name=", "leaving_date=", "file", "initial_date="])
    except getopt.GetoptError as err:
        print(err)
        sys.exit(2)
    
    command = Command.initial
    for o, a in opts:
        if o in ("-m", "--member_name"):
            app_params[const.APP_PARAMS_MEMBER_NAME] = a
        elif o in ("-l", "--leaving_date"):
            app_params[const.APP_PARAMS_LEAVING_DATE] = a
        elif o in ("-f", "--file"):
            app_params[const.APP_PARAMS_FILE_NAME] = a
        elif o in ("-d", "--initial_date"):
            app_params[const.APP_PARAMS_INITIAL_DATE] = a
        elif o in ("-a"):
            command = Command.addMember
        elif o in ("-i"):
            command = Command.initial
        elif o in ("-r"):
            command = Command.recordLeaving
    
    

    file_name = app_params[const.APP_PARAMS_FILE_NAME]
    if file_name == None or file_name == '':
        print("The attentance file can't not be empty, assigning it by using '-f' or '--file-name'")
        sys.exit(2)
    
    if os.path.exists(file_name):
        book = openpyxl.load_workbook(file_name)
    else:
        book = openpyxl.Workbook(write_only=True)
        book.save(file_name)
    try:
        if command == Command.addMember:
            perform_add_member(book, app_params)
        elif command == Command.initial:
            perform_initial_attendance(book, app_params)
        elif command == Command.recordLeaving:
            perform_record_leaving(book, app_params)
        
    except ValueError as err:
        print(err)
        sys.exit(2)




if __name__ == "__main__":
    main()