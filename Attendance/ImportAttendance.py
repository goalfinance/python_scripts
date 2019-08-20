from getopt import getopt
from getopt import GetoptError
import sys
from goalfinance.utils.utils import Const
from excel.translator import (load_workbook, translate)
import time
import openpyxl

app_params = dict()
const = Const()
const.APP_PARAMS_SOURCE_FILE = "source"
const.APP_PARAMS_TARGET_FILE = "target"
const.APP_PARAMS_MONTH = "month"

def getMembersAttendance(workbook):
    pass

def main():
    try:
        opts, args = getopt(sys.argv[1:], "s:t:m:", [const.APP_PARAMS_SOURCE_FILE, const.APP_PARAMS_TARGET_FILE, const.APP_PARAMS_MONTH])
    except GetoptError as opt_error:
        print(opt_error)
        exit(2)

    for o, a in opts:
        if o in ("-s", "--source"):
            app_params[const.APP_PARAMS_SOURCE_FILE] = a
        elif o in ("-t", "--target"):
            app_params[const.APP_PARAMS_TARGET_FILE] = a
        elif o in ("-m", "--month"):
            app_params[const.APP_PARAMS_MONTH] = a
    
    source_workbook = load_workbook(app_params[const.APP_PARAMS_SOURCE_FILE])
    if  source_workbook == None:
        print("The source file should not be absent, please assigning it by using '-s' or '--source'")
        sys.exit(2)
    

    if app_params[const.APP_PARAMS_SOURCE_FILE] == None or app_params[const.APP_PARAMS_SOURCE_FILE] == "":
        print("The target file is absent, so use the default[target file name = '.Attendance statistics of cassiopae.xlsx']")
        app_params[const.APP_PARAMS_TARGET_FILE] = ".Attendance statistics of cassiopae.xlsx"

    target_workbook = load_workbook(app_params[const.APP_PARAMS_TARGET_FILE])
    if target_workbook == None:
        target_workbook = openpyxl.Workbook(write_only=True)
        target_workbook.save(app_params[const.APP_PARAMS_TARGET_FILE])
        
    trans_date = None
    if app_params[const.APP_PARAMS_MONTH] == None:
        print("The month of attendance should not be absent, please assigning it by using '-m' or '--month', use the format '%Y-%m', for example, '2019-05'")
        sys.exit(2)
    elif app_params[const.APP_PARAMS_MONTH] != "":
        try:
            trans_date = time.strptime(app_params[const.APP_PARAMS_MONTH], "%Y-%m")
        except ValueError as err:
            raise IllegalParamsError("The value of '" + const.APP_PARAMS_MONTH + "' is illegal, the correct format is '%Y-%m', for example, '2019-05'", err)
    
    translate(source_workbook, target_workbook, trans_date)

    target_workbook.save(app_params[const.APP_PARAMS_TARGET_FILE])
    
if __name__ == "__main__":
    main()


    

    

     




